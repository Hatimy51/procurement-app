"""
Converts non-text enquiry sources (Excel files, screenshots) into plain text,
so they can flow through the SAME extraction engine as pasted email text.
This keeps one extraction pipeline instead of building separate logic per
file format — the only thing that changes per format is how we get to text.
"""
import io


def read_xlsx_as_text(file_bytes: bytes) -> str:
    """
    Reads every sheet/row/cell of an Excel file and renders it as a plain
    text table. Excel enquiries are already structured, so this text is
    usually cleaner and easier to extract from than free-form email prose.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    lines = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):  # skip fully empty rows
                lines.append(" | ".join(cells))
    return "\n".join(lines)


def read_csv_as_text(file_bytes: bytes) -> str:
    return file_bytes.decode("utf-8", errors="ignore")


def read_image_as_text(file_bytes: bytes) -> str:
    """
    OCR for screenshots/photos of enquiries (e.g. a tabled enquiry embedded
    in an email as an image, which can't be copy-pasted as text at all).
    Uses Tesseract — free, runs locally, no per-use cost, consistent with
    the zero-cost v1 constraint. Expect this to be the least reliable of
    the three input paths (OCR quality varies with image clarity).
    """
    import pytesseract
    from PIL import Image

    image = Image.open(io.BytesIO(file_bytes))
    return pytesseract.image_to_string(image)


def read_pdf_as_text(file_bytes: bytes) -> str:
    """
    Extracts text from a PDF's actual text layer (works for quotes generated
    directly as PDFs — the common case for a supplier's formal quotation).
    Does NOT do OCR — a scanned/photographed PDF with no real text layer
    will come back empty, and callers should treat an empty result as
    "this PDF has no readable text" rather than "this PDF has no items."
    """
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(file_bytes))
    pages_text = [page.extract_text() or "" for page in reader.pages]
    return "\n".join(pages_text).strip()


def parse_tabular_file(filename: str, file_bytes: bytes) -> tuple[list[str], list[list[str]]]:
    """
    Reads an Excel/CSV file as structured rows (headers + data rows), for
    the Import screen — where we need actual columns to map, unlike
    read_xlsx_as_text() above which flattens everything for the enquiry
    extraction pipeline.
    """
    lower = filename.lower()
    if lower.endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], []
        headers = [str(c).strip() if c is not None else "" for c in header_row]
        rows = []
        for row in rows_iter:
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                rows.append(cells)
        return headers, rows

    elif lower.endswith(".csv"):
        import csv

        text = file_bytes.decode("utf-8", errors="ignore")
        reader = list(csv.reader(io.StringIO(text)))
        if not reader:
            return [], []
        headers = [h.strip() for h in reader[0]]
        rows = [r for r in reader[1:] if any(c.strip() for c in r)]
        return headers, rows

    else:
        raise ValueError(
            f"Unsupported file type for import: '{filename}'. "
            "Supported for import: .xlsx, .xls, .csv"
        )

def extract_text_from_upload(filename: str, file_bytes: bytes) -> str:
    """Dispatches to the right reader based on file extension."""
    lower = filename.lower()
    if lower.endswith((".xlsx", ".xls")):
        return read_xlsx_as_text(file_bytes)
    elif lower.endswith(".csv"):
        return read_csv_as_text(file_bytes)
    elif lower.endswith((".png", ".jpg", ".jpeg", ".webp", ".gif")):
        return read_image_as_text(file_bytes)
    elif lower.endswith(".pdf"):
        text = read_pdf_as_text(file_bytes)
        if not text:
            raise ValueError(
                f"'{filename}' has no readable text — it's likely a scanned/photographed "
                "PDF rather than one generated digitally. OCR for scanned PDFs isn't "
                "supported yet; try re-saving it as an image (screenshot) instead, which "
                "does go through OCR."
            )
        return text
    else:
        raise ValueError(
            f"Unsupported file type for '{filename}'. "
            "Supported: .xlsx, .xls, .csv, .pdf, .png, .jpg, .jpeg, .webp"
        )
